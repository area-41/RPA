import pdfplumber
import re
import shutil
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers, Border, Side

# Extrai o CNPJ usando RegEx
def extrair_cnpj(texto):
  resultado = re.search(r'(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})', texto)
  if resultado:
    return resultado.group(1)
  return ''

# Define o valor e aplica estilos em uma célula
def cabecalho_valor_estilo(aba, linha, coluna, valor, estilo):
  celula = aba.cell(row=linha, column=coluna, value=valor)
  celula.fill = estilo['fill']
  celula.font = estilo['font']
  celula.alignment = estilo['alignment']

#########################################################
# Extração dos dados dos arquivos pdf das notas fiscais
#########################################################

# Extração de todas as notas para o diretório Notas
arquivo_notas_fiscais = 'Notas_Fiscais.zip'
diretorio_notas = 'Notas'
shutil.unpack_archive(arquivo_notas_fiscais, diretorio_notas)

caminho_diretorio = Path(diretorio_notas)

# Leitura de cada um dos pdfs no diretório Notas
dados_notas = []
for arquivo_pdf in caminho_diretorio.glob('*.pdf'):
  with pdfplumber.open(arquivo_pdf) as pdf:
    # Extração do texto completo e tabelas do pdf
    texto_completo = ''
    tabelas = []
    for pagina in pdf.pages:
      texto = pagina.extract_text()
      tabelas_pagina = pagina.extract_tables()
      if texto:
        texto_completo += texto + '\n'
      if tabelas_pagina:
        for tabela in tabelas_pagina:
          tabelas.append(tabela)

    # Início das extrações das informações
    # Extrair número da nota fiscal
    resultado_numero = re.search(r'NOTA FISCAL N. (\d+)', texto_completo, re.IGNORECASE)
    numero_nota = resultado_numero.group(1) if resultado_numero else ''

    # Extrair data de emissão
    resultado_data = re.search(r'DATA DE EMISSÃO[:\s]+ (\d{2}/\d{2}/\d{4})', texto_completo, re.IGNORECASE)
    data_emissao = resultado_data.group(1) if resultado_data else ''

    # Extrair informações do emitente
    emitente = {}
    resultado_emitente = re.search(r'EMITENTE[\s]*[Razão Social]+[:\s]+(.*)[\s]*[CNPJ]+(.*)', texto_completo, re.IGNORECASE)
    emitente['razao_social'] = resultado_emitente.group(1) if resultado_emitente else ''
    emitente['cnpj'] = extrair_cnpj(resultado_emitente.group(2)) if resultado_emitente else ''

    # Extrair informações do destinatário
    destinatario = {}
    resultado_destinatario = re.search(r'DESTINATÁRIO[\s]*[Razão Social]+[:\s]+(.*)[\s]*[CNPJ]+(.*)', texto_completo, re.IGNORECASE)
    destinatario['razao_social'] = resultado_destinatario.group(1) if resultado_destinatario else ''
    destinatario['cnpj'] = extrair_cnpj(resultado_destinatario.group(2)) if resultado_destinatario else ''

    # Extrair informações dos produtos
    produtos = []

    # Extrair tabela de produtos
    for tabela in tabelas:
      # Verificar se a palavra código ou descrição está na primeira linha da tabela
      if 'Código'.upper() in str(tabela[0]).upper() or 'Descrição'.upper() in str(tabela[0]).upper():
        for linha in tabela[1:-1]:
          resultado_valor_unitario = re.search(r'(\d+[.,]\d{2})', linha[4])
          valor_unitario_float = float(resultado_valor_unitario.group(1)) if resultado_valor_unitario else 0.0
          resultado_valor_total = re.search(r'(\d+[.,]\d{2})', linha[5])
          valor_total_float = float(resultado_valor_total.group(1)) if resultado_valor_total else 0.0
          produto = {
            'codigo': linha[0],
            'descricao': linha[1],
            'quantidade': int(linha[2]),
            'unidade': linha[3],
            'valor_unitario': valor_unitario_float,
            'valor_total': valor_total_float,
          }
          produtos.append(produto)

    # Extrair total dos valores de produtos
    resultado_total_produtos = re.search(r'TOTAL[:\s]+[R\$ ]+(\d+[.,]?\d{2})', texto_completo, re.IGNORECASE)
    total_produtos = float(resultado_total_produtos.group(1)) if resultado_total_produtos else 0.0

    # Extrair total dos impostos
    resultado_total_impostos = re.search(r'IMPOSTO[S]?[:\s]+[R\$ ]+(\d+[.,]?\d{2})', texto_completo, re.IGNORECASE)
    total_impostos = float(resultado_total_impostos.group(1)) if resultado_total_impostos else 0.0

    # Armazena todos os dados em um dicionário e adiciona na lista
    dados_notas.append({
        'numero': numero_nota,
        'data_emissao': data_emissao,
        'emitente': emitente,
        'destinatario': destinatario,
        'produtos': produtos,
        'total_produtos': total_produtos,
        'total_impostos': total_impostos
    })

################################################################
# Criação de um arquivo xlsx para armazenar os dados extraídos
################################################################

arquivo_relatorio = 'relatorio_notas_fiscais.xlsx'
wb_relatorio = Workbook()

aba_resumo = wb_relatorio.active
aba_resumo.title = 'Resumo das Notas Fiscais'

# Cabeçalho da aba resumo
cabecalho_resumo = [
  'Número',
  'Data Emissão',
  'Emitente',
  'CNPJ Emitente',
  'Destinatário',
  'CNPJ Destinatário',
  'Total Produtos (R$)',
  'Total Impostos (R$)',
  'Total Nota (R$)',
  'Quantidade Itens'
]

# Escrever cabeçalho e definir estilos
estilo_cabecalho = {
  'font': Font(bold=True, color='FFFFFF'),
  'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
  'alignment': Alignment(horizontal='center', vertical='center', wrapText=True),
}
for coluna, titulo in enumerate(cabecalho_resumo, start=1):
  cabecalho_valor_estilo(aba_resumo, 1, coluna, titulo, estilo_cabecalho)

# Escrever dados do resumo
for i, nota in enumerate(dados_notas, start=2):
  aba_resumo.cell(row=i, column=1, value=nota['numero'])
  aba_resumo.cell(row=i, column=2, value=nota['data_emissao'])
  celula_razao_social_e = aba_resumo.cell(row=i, column=3, value=nota['emitente']['razao_social'])
  aba_resumo.cell(row=i, column=4, value=nota['emitente']['cnpj'])
  celula_razao_social_d = aba_resumo.cell(row=i, column=5, value=nota['destinatario']['razao_social'])
  aba_resumo.cell(row=i, column=6, value=nota['destinatario']['cnpj'])
  celula_total_produtos = aba_resumo.cell(row=i, column=7, value=nota['total_produtos'])
  celula_total_impostos = aba_resumo.cell(row=i, column=8, value=nota['total_impostos'])
  celula_total_nota = aba_resumo.cell(row=i, column=9, value=f'=SUM(G{i},H{i})')
  aba_resumo.cell(row=i, column=10, value=len(nota['produtos']))

  # Definição de formatação e alinhamento
  celula_razao_social_e.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
  celula_razao_social_d.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
  celula_total_produtos.number_format = numbers.FORMAT_NUMBER_00
  celula_total_produtos.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
  celula_total_impostos.number_format = numbers.FORMAT_NUMBER_00
  celula_total_impostos.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
  celula_total_nota.number_format = numbers.FORMAT_NUMBER_00
  celula_total_nota.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

# Totais gerais de todas as notas fiscais
ultima_linha = aba_resumo.max_row
linha_totais = ultima_linha+2
aba_resumo.merge_cells(start_row=linha_totais, end_row=linha_totais, start_column=6, end_column=7)
cabecalho_valor_estilo(aba_resumo, linha_totais, 6, 'TOTAIS', estilo_cabecalho)
cabecalho_valor_estilo(aba_resumo, linha_totais+1, 6, 'Total em produtos:', estilo_cabecalho)
celula_total_produtos = aba_resumo.cell(row=linha_totais+1, column=7, value=f'=SUM(G2:G{ultima_linha})')
celula_total_produtos.font = Font(bold=True, color='000000')
cabecalho_valor_estilo(aba_resumo, linha_totais+2, 6, 'Total em impostos:', estilo_cabecalho)
celula_total_impostos = aba_resumo.cell(row=linha_totais+2, column=7, value=f'=SUM(H2:H{ultima_linha})')
celula_total_impostos.font = Font(bold=True, color='000000')
cabecalho_valor_estilo(aba_resumo, linha_totais+3, 6, 'Total geral:', estilo_cabecalho)
celula_total = aba_resumo.cell(row=linha_totais+3, column=7, value=f'=SUM(I2:I{ultima_linha})')
celula_total.font = Font(bold=True, color='000000')

# Definir largura de cada coluna
aba_resumo.column_dimensions['A'].width = 10
aba_resumo.column_dimensions['B'].width = 15
aba_resumo.column_dimensions['C'].width = 20
aba_resumo.column_dimensions['D'].width = 20
aba_resumo.column_dimensions['E'].width = 20
aba_resumo.column_dimensions['F'].width = 20
aba_resumo.column_dimensions['G'].width = 12
aba_resumo.column_dimensions['H'].width = 12
aba_resumo.column_dimensions['I'].width = 12
aba_resumo.column_dimensions['J'].width = 10

# Criar aba de produtos
aba_produtos = wb_relatorio.create_sheet(title='Produtos', index=1)

# Cabeçalho da aba Produtos
cabecalho_produtos = [
  'Nº NF',
  'Data Emissão',
  'Destinatário',
  'Cód. Produto',
  'Descrição',
  'Unidade',
  'Quantidade',
  'Valor Unit. (R$)',
  'Valor Total (R$)'
]

# Escrever cabeçalho e definir estilos
for coluna, titulo in enumerate(cabecalho_produtos, start=1):
  cabecalho_valor_estilo(aba_produtos, 1, coluna, titulo, estilo_cabecalho)

# Escrever todos os produtos de cada nota
linha_produto = 2
for nota in dados_notas:
  for produto in nota['produtos']:
    aba_produtos.cell(row=linha_produto, column=1, value=nota['numero'])
    aba_produtos.cell(row=linha_produto, column=2, value=nota['data_emissao'])
    celula_razao_social_d = aba_produtos.cell(row=linha_produto, column=3, value=nota['destinatario']['razao_social'])
    aba_produtos.cell(row=linha_produto, column=4, value=produto['codigo'])
    celula_descricao = aba_produtos.cell(row=linha_produto, column=5, value=produto['descricao'])
    aba_produtos.cell(row=linha_produto, column=6, value=produto['unidade'])
    celula_quantidade = aba_produtos.cell(row=linha_produto, column=7, value=produto['quantidade'])
    celula_valor_unitario = aba_produtos.cell(row=linha_produto, column=8, value=produto['valor_unitario'])
    celula_valor_total = aba_produtos.cell(row=linha_produto, column=9, value=produto['valor_total'])

    # Definição de formatação e alinhamento
    celula_descricao.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    celula_razao_social_d.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
    celula_quantidade.number_format = '0'
    celula_quantidade.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    celula_valor_unitario.number_format = numbers.FORMAT_NUMBER_00
    celula_valor_unitario.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    celula_valor_total.number_format = numbers.FORMAT_NUMBER_00
    celula_valor_total.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

    linha_produto += 1

# Definir largura de cada coluna
aba_produtos.column_dimensions['A'].width = 10
aba_produtos.column_dimensions['B'].width = 15
aba_produtos.column_dimensions['C'].width = 20
aba_produtos.column_dimensions['D'].width = 10
aba_produtos.column_dimensions['E'].width = 20
aba_produtos.column_dimensions['F'].width = 10
aba_produtos.column_dimensions['G'].width = 12
aba_produtos.column_dimensions['H'].width = 12
aba_produtos.column_dimensions['I'].width = 12

# Salva o arquivo final
wb_relatorio.save(arquivo_relatorio)
print(f'Arquivo {arquivo_relatorio} criado!')
